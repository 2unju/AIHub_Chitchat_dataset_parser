import os
import re
import random
import pandas as pd
from transformers import PreTrainedTokenizerFast
from tqdm import tqdm
from openpyxl import Workbook
from sklearn.model_selection import train_test_split

from arguments import get_args


def save(result):
    with open("dialogue_chatbot.tsv", "a", encoding="utf-8") as fp:
        for line in tqdm(result, "[SAVE]"):
            fp.write(f"{line[0]}\t{line[1]}\n")

def make(args):
    tok = PreTrainedTokenizerFast.from_pretrained(args.tok)
    private = re.compile("#[가-힣]+#")

    dirlist = [dirname for dirname in os.listdir(args.data_path) if dirname.endswith(".xlsx")]
    j = [dirname for dirname in os.listdir(args.data_path) if not dirname.endswith(".xlsx")
         and not dirname.endswith(".zip")][0]
    for dirname in dirlist:
        result = []
        skip = False

        path = os.path.join(os.path.dirname(__file__), args.data_path, dirname)
        data = pd.read_excel(path, index_col=None)

        temp = []
        predomain = "first"
        qdomain = "first"
        for sentence, id, swt, domain in tqdm(zip(data["SENTENCE"], data["SENTENCEID"], data["QA"], data["MAIN"]),
                                            f"{dirname}", total=len(data)):
            if swt == "Q":
                qdomain = domain
            if id == 1:
                skip = False
                temp = []
                qdomain = "first"
            if (not temp and swt == "A") or (temp and swt == "Q"):
                temp = []
                continue
            temp.append(sentence)

            if len(tok.tokenize(sentence)) <= 3:
                skip = True
            if len(temp) == 2:
                if not skip and (qdomain in ["first", domain] or qdomain != predomain)\
                        and not (private.search(temp[0]) or private.search(temp[1])):
                    result.append(temp)
                skip = False
                temp = []
            predomain = domain
        save(result)

    dirlist = [dirname for dirname in os.listdir(os.path.join(os.path.dirname(__file__), args.data_path, j))
               if dirname.endswith(".xlsx")]
    for dirname in dirlist:
        path = os.path.join(os.path.dirname(__file__), args.data_path, j, dirname)
        data = pd.read_excel(path, index_col=None)
        data.dropna(subset=["question", "answer"], inplace=True)

        result = []
        for q, a in tqdm(zip(data["question"], data["answer"]), f"{dirname}",
                         total=len(data)):
            if len(tok.tokenize(q)) <= 3 or len(tok.tokenize(a)) <= 3:
                continue
            result.append([q, a])
        save(result)


def sample():
    with open("dialogue_chatbot.tsv", "r", encoding="utf-8") as fp:
        data = fp.read()
        data = data.split("\n")
    data = data[:-1]
    for i in range(len(data)):
        data[i] = data[i].split("\t")
    print(f"with duplication : {len(data)}")
    data = list(set([tuple(item) for item in data]))
    print(f"without duplication : {len(data)}")
    random.shuffle(data)
    sampledata = data[:100]

    wb = Workbook()
    sheet = wb.active

    for id, (q, a) in enumerate(sampledata):
        sheet.cell(row=id+1, column=1).value = q
        sheet.cell(row=id+1, column=2).value = a

    wb.save(filename="sample1.xlsx")


def split():
    with open("dialogue_chatbot.tsv", "r", encoding="utf-8") as fp:
        data = fp.read()
        data = data.split("\n")
    data = data[:-1]
    for i in range(len(data)):
        data[i] = data[i].split("\t")
    train, valid = train_test_split(data, test_size=0.1, shuffle=True)
    valid, test = train_test_split(valid, test_size=0.5, shuffle=True)

    if not os.path.exists("result/"):
        os.mkdir("result")

    with open("result/train.tsv", "w", encoding="utf-8") as fp:
        for q, r in train:
            fp.write(f"{q}\t{r}\n")
    with open("result/valid.tsv", "w", encoding="utf-8") as fp:
        for q, r in valid:
            fp.write(f"{q}\t{r}\n")
    with open("result/test.tsv", "w", encoding="utf-8") as fp:
        for q, r in test:
            fp.write(f"{q}\t{r}\n")


if __name__ == "__main__":
    args = get_args()
    if args.mode == "all":
        make(args)
        split()
        sample()
    elif args.mode == "make":
        make(args)
        split()
    else:
        sample()
